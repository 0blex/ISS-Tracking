# -*- coding: utf-8 -*-
"""
Created on Tue May 31 20:25:58 2022

@author: AlexBlack
"""

# -*- coding: utf-8 -*-
"""
Created on Sat Feb 12 20:36:15 2022

@author: AlexBlack
"""

import time
from openpyxl import Workbook,load_workbook
import pandas as pd
from datetime import datetime
from threading import Thread, Event
from functions.functions import log
from config.directories import rootDir


intervals = int(input('How frequently should data be collected (in seconds): '))
length = int(input('How long should data be collected for (in seconds): '))
datenow = str(datetime.now().strftime("%Y-%m-%d"))
timenow = str(datetime.now().strftime("%H-%M-%S"))


log('Program started')


# create workbook to record results
wb = Workbook()
wb.active.title = 'results'
wb.create_sheet('readme')

wb['readme']['B2'] = 'Positional data of the ISS recorded at {} at {} for {} seconds in intervals of {} seconds.'.format(datenow,timenow,length,intervals)
wb['results'].append(['timestamp','latitude','longitude'])

sheetname = datenow+" "+str(intervals)+"-"+str(length)
wb.save(rootDir+'\output\\{0}.xlsx'.format(sheetname))


data = pd.DataFrame()
ws = wb['results']
url = 'http://api.open-notify.org/iss-now.json'


#begin collecting data
stop = Event()

def read_api():
    while not stop.is_set():
        global data 
        df = pd.read_json(url)
        
        df['latitude'] = df.loc['latitude','iss_position']
        df['longitude'] = df.loc['longitude','iss_position']
        df.reset_index(inplace=True)
        df = df.drop(['iss_position','index','message'], axis=1).loc[:0]
        
        data = pd.concat([data,df])
        
        row = list(df.iloc[0])
        ws.append(row)
        wb.save(rootDir+'\output\\{0}.xlsx'.format(sheetname))

        log('Successful data collection')
            
        time.sleep(intervals)

thread = Thread(target=read_api)

thread.start()
thread.join(timeout=length)
stop.set()

file = rootDir+'\output\{} series.xlsx'.format(sheetname)
data.to_excel(file,index=False)

log('Program completed')


